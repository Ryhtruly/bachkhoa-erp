import os
import shutil
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter

EXCEL_PATH = os.path.join(os.getcwd(), "tai_lieu", "co_so_du_lieu", "BACHKHOA_FULL_AUTO_2026_migrated.xlsx")
BACKUP_DIR = os.path.join(os.getcwd(), "backups")

def backup_database():
    """Create a backup of the Excel master database."""
    os.makedirs(BACKUP_DIR, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = os.path.join(BACKUP_DIR, f"master_backup_{timestamp}.xlsx")
    try:
        shutil.copy2(EXCEL_PATH, backup_path)
        # Keep only the last 10 backups to save space
        backups = sorted([os.path.join(BACKUP_DIR, f) for f in os.listdir(BACKUP_DIR) if f.startswith("master_backup_")])
        if len(backups) > 10:
            for b in backups[:-10]:
                os.remove(b)
        return backup_path
    except Exception as e:
        print(f"Error creating backup: {e}")
        return None

def read_sheet(sheet_name, data_only=True):
    """Read a sheet from the Excel file and return a list of dictionaries."""
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=data_only)
        if sheet_name not in wb.sheetnames:
            return []
        
        sheet = wb[sheet_name]
        
        # In Full Auto file, usually Row 3 or Row 1 has headers depending on the sheet.
        # Let's find the header row by looking for known column names, or default to row 3 (standard header in DATABASE_HOSO).
        header_row = 3
        if sheet_name == "THIET_LAP":
            header_row = 3
        elif sheet_name == "MAP_NGUON_DU_LIEU" or sheet_name == "Sheet1":
            header_row = 1
            
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
            
        headers = [str(cell).strip() if cell is not None else None for cell in rows[header_row - 1]]
        
        # Check if headers are empty, if so find the first non-empty row as header
        if not any(headers):
            for idx, r in enumerate(rows):
                non_none = [c for c in r if c is not None]
                if len(non_none) > 3:
                    headers = [str(cell).strip() if cell is not None else None for cell in r]
                    header_row = idx + 1
                    break
        
        # Keep None headers to preserve column alignment, but skip them when building dict
        num_cols = len(headers)
        
        data = []
        for r in rows[header_row:]:
            if not any(r[:num_cols]): # Empty row
                continue
            row_dict = {}
            for col_idx, h in enumerate(headers):
                if h is None:
                    continue
                val = r[col_idx]
                # Format dates and None
                if isinstance(val, datetime):
                    val = val.strftime("%Y-%m-%d")
                row_dict[h] = val
            data.append(row_dict)
            
        wb.close()
        return data
    except Exception as e:
        print(f"Error reading sheet {sheet_name}: {e}")
        return []

def get_next_id(sheet_name, id_prefix, col_idx=1):
    """Get the next ID like BK-HS-0321."""
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        sheet = wb[sheet_name]
        
        header_row = 3
        max_num = 0
        
        for row in range(header_row + 1, sheet.max_row + 1):
            val = sheet.cell(row=row, column=col_idx).value
            if val and str(val).startswith(id_prefix):
                try:
                    num_str = str(val).replace(id_prefix, "")
                    num = int(num_str)
                    if num > max_num:
                        max_num = num
                except ValueError:
                    continue
        wb.close()
        return f"{id_prefix}{str(max_num + 1).zfill(4)}"
    except Exception as e:
        print(f"Error generating next ID for {sheet_name}: {e}")
        return f"{id_prefix}0001"

def write_to_sheet(sheet_name, new_row_data, header_row=3):
    """Safely append a new row of data to the specified sheet."""
    backup_database()
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=False)
        if sheet_name not in wb.sheetnames:
            return False, f"Sheet {sheet_name} not found."
            
        sheet = wb[sheet_name]
        
        # Get headers
        headers = [str(sheet.cell(row=header_row, column=col).value).strip() 
                   for col in range(1, sheet.max_column + 1) 
                   if sheet.cell(row=header_row, column=col).value is not None]
        
        # Find next empty row (look for first row with empty ID column)
        insert_row = sheet.max_row + 1
        for row in range(header_row + 1, sheet.max_row + 2):
            cell_val = sheet.cell(row=row, column=1).value
            if cell_val is None or str(cell_val).strip() == "":
                insert_row = row
                break
                
        # Write values
        for col_idx, header in enumerate(headers, 1):
            if header in new_row_data:
                val = new_row_data[header]
                # Handle formula formatting
                if isinstance(val, str) and val.startswith("="):
                    # Replace placeholders like {row} with actual row index
                    val = val.replace("{row}", str(insert_row))
                sheet.cell(row=insert_row, column=col_idx, value=val)
                
        wb.save(EXCEL_PATH)
        wb.close()
        return True, "Success"
    except Exception as e:
        print(f"Error writing to sheet {sheet_name}: {e}")
        return False, str(e)

def update_row_in_sheet(sheet_name, key_col_name, key_val, updates, header_row=3):
    """Update cells in an existing row of data matching key_val."""
    backup_database()
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=False)
        if sheet_name not in wb.sheetnames:
            return False, f"Sheet {sheet_name} not found."
            
        sheet = wb[sheet_name]
        
        # Find key column index
        key_col_idx = -1
        headers = {}
        for col in range(1, sheet.max_column + 1):
            h_val = sheet.cell(row=header_row, column=col).value
            if h_val is not None:
                h_name = str(h_val).strip()
                headers[h_name] = col
                if h_name == key_col_name:
                    key_col_idx = col
                    
        if key_col_idx == -1:
            return False, f"Key column {key_col_name} not found."
            
        # Find the row
        found_row = -1
        for row in range(header_row + 1, sheet.max_row + 1):
            cell_val = sheet.cell(row=row, column=key_col_idx).value
            if cell_val and str(cell_val).strip() == str(key_val).strip():
                found_row = row
                break
                
        if found_row == -1:
            return False, f"Row with {key_col_name}={key_val} not found."
            
        # Apply updates
        for h_name, val in updates.items():
            if h_name in headers:
                col_idx = headers[h_name]
                if isinstance(val, str) and val.startswith("="):
                    val = val.replace("{row}", str(found_row))
                sheet.cell(row=found_row, column=col_idx, value=val)
                
        wb.save(EXCEL_PATH)
        wb.close()
        return True, "Success"
    except Exception as e:
        print(f"Error updating sheet {sheet_name}: {e}")
        return False, str(e)

# Specific operations

def get_dashboard_summary():
    """Calculate and return dashboard KPIs."""
    hoso = read_sheet("DATABASE_HOSO")
    hopdong = read_sheet("DATABASE_HOP_DONG")
    
    total_hoso = len([h for h in hoso if h.get("Mã hồ sơ")])
    in_progress = len([h for h in hoso if h.get("Mã hồ sơ") and h.get("Trạng thái") not in ["Hoàn thành", "Đã hủy", "Nộp thành công - Chờ kết quả"]])
    completed = len([h for h in hoso if h.get("Mã hồ sơ") and h.get("Trạng thái") in ["Hoàn thành", "Nộp thành công - Chờ kết quả"]])
    
    overdue = 0
    for h in hoso:
        if h.get("Mã hồ sơ") and h.get("Trạng thái") not in ["Hoàn thành", "Đã hủy"]:
            # If Cảnh báo contains 'QUÁ HẠN'
            cb = str(h.get("Cảnh báo") or "")
            if "QUÁ HẠN" in cb.upper() or (h.get("Số ngày còn lại") is not None and isinstance(h.get("Số ngày còn lại"), (int, float)) and h.get("Số ngày còn lại") < 0):
                overdue += 1
                
    total_val = 0
    total_collected = 0
    for hd in hopdong:
        if hd.get("Mã hợp đồng"):
            val = hd.get("Giá trị hợp đồng") or hd.get("Giá trị HĐ") or 0
            col = hd.get("Đã thu") or hd.get("Đã thanh toán") or 0
            try:
                total_val += float(val)
                total_collected += float(col)
            except ValueError:
                continue
                
    debt = total_val - total_collected
    
    recent_hoso = hoso[-10:] if len(hoso) > 10 else hoso
    recent_hoso.reverse()
    
    return {
        "total_hoso": total_hoso,
        "in_progress": in_progress,
        "completed": completed,
        "overdue": overdue,
        "total_contract_value": total_val,
        "total_collected": total_collected,
        "total_debt": debt,
        "recent_hoso": recent_hoso[:10]
    }

def add_new_hoso(hoso_data):
    """Add a new profile and auto-generate its code."""
    next_id = get_next_id("DATABASE_HOSO", "BK-HS-")
    hoso_data["Mã hồ sơ"] = next_id
    hoso_data["Ngày tạo"] = datetime.now().strftime("%Y-%m-%d")
    hoso_data["Phòng ban"] = "Phòng Đo đạc"
    hoso_data["Số ngày còn lại"] = "=IF(M{row}=\"\",\"\",M{row}-TODAY())"
    hoso_data["Cảnh báo"] = "=IF(A{row}=\"\",\"\",IF(L{row}=\"Hoàn thành\",\"XONG\",IF(N{row}<0,\"QUÁ HẠN\",\"Trong hạn\")))"
    
    # Write to DATABASE_HOSO
    success, err = write_to_sheet("DATABASE_HOSO", hoso_data)
    if not success:
        return False, err
        
    # Automatically sync to DODAC_TAC_NGHIEP
    tac_nghiep_data = {
        "Mã hồ sơ": next_id,
        "Khách": hoso_data.get("Tên khách hàng"),
        "Địa chỉ/Phường": hoso_data.get("Khu vực/Phường"),
        "Loại đo": hoso_data.get("Loại dịch vụ"),
        "Nhân viên đo": hoso_data.get("Phụ trách chính"),
        "Phụ đo": hoso_data.get("Hỗ trợ"),
        "Ngày đo": hoso_data.get("Ngày tạo"),
        "Kết quả hiện trường": "Trong hạn",
        "Trạng thái đo": "Trong hạn",
        "Phụ cấp": 0,
        "Ghi chú": ""
    }
    write_to_sheet("DODAC_TAC_NGHIEP", tac_nghiep_data)
    
    # Create matching task in DATABASE_TASK
    task_data = {
        "Mã task": get_next_id("DATABASE_TASK", "BK-TSK-"),
        "Mã hồ sơ": next_id,
        "Phòng ban": "Phòng Đo đạc",
        "Tên công việc": hoso_data.get("Loại dịch vụ"),
        "Người giao": "Giám đốc",
        "Người nhận": hoso_data.get("Phụ trách chính"),
        "Ngày giao": hoso_data["Ngày tạo"],
        "Deadline": hoso_data.get("Deadline"),
        "Trạng thái": "Trong hạn",
        "Kết quả": "Chờ triển khai",
        "Quá hạn?": '=IF(OR(H{row}="",J{row}="Hoàn thành",J{row}="Đã hủy"),"Không",IF(H{row}<TODAY(),"Có","Không"))',
        "Ghi chú": ""
    }
    write_to_sheet("DATABASE_TASK", task_data)
    
    return True, next_id

def add_new_hopdong(hd_data):
    """Add a new contract."""
    # Write to DATABASE_HOP_DONG
    hd_data["Còn nợ"] = "=IF(G{row}=\"\",\"\",G{row}-H{row})"
    hd_data["Tình trạng"] = "=IF(A{row}=\"\",\"\",IF(I{row}<=0,\"Đã tất toán\",IF(L{row}<TODAY(),\"Quá hạn thanh toán\",\"Chờ thanh toán\")))"
    
    success, err = write_to_sheet("DATABASE_HOP_DONG", hd_data)
    if not success:
        return False, err
        
    return True, hd_data["Mã hợp đồng"]

def add_new_thuchi(tc_data):
    """Add an income/expense record."""
    tc_data["Ngày"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    tc_data["Thu (+)"] = "=IF(C{row}=\"Thu\",J{row},0)"
    tc_data["Chi (-)"] = "=IF(C{row}=\"Chi\",J{row},0)"
    tc_data["Ghi chú"] = "=K{row}-L{row}" # Net change
    
    success, err = write_to_sheet("DATABASE_THU_CHI", tc_data)
    return success, err

def calculate_luong_khoan_report(month_str):
    """
    Calculate subcontract salary (Lương khoán 3P) based on Bảng 02. Cơ chế khoán đo vẽ.xlsx 
    and DATABASE_HOSO completed items.
    """
    hoso = read_sheet("DATABASE_HOSO")
    # Read contract rates from Bảng 02
    rate_path = os.path.join(os.getcwd(), "tai_lieu", "co_so_du_lieu", "Bảng 02. Cơ chế khoán đo vẽ.xlsx")
    rate_wb = openpyxl.load_workbook(rate_path, data_only=True)
    rate_sheet = rate_wb.active
    rates = {}
    for r in range(2, rate_sheet.max_row + 1):
        service = rate_sheet.cell(row=r, column=1).value
        main_pay = rate_sheet.cell(row=r, column=2).value or 0
        assistant_pay = rate_sheet.cell(row=r, column=3).value or 0
        if service:
            rates[str(service).strip()] = {
                "main": float(main_pay),
                "assistant": float(assistant_pay)
            }
    rate_wb.close()
    
    # Calculate for each completed profile
    payroll_records = []
    
    for h in hoso:
        if h.get("Mã hồ sơ") and h.get("Trạng thái") == "Hoàn thành":
            # Format: 'BK-HS-xxxx'
            # Let's check date created or deadline for month matching
            created_date = h.get("Ngày tạo") or ""
            if created_date.startswith(month_str):
                service = str(h.get("Loại dịch vụ") or "").strip()
                main_staff = h.get("Phụ trách chính")
                assist_staff = h.get("Hỗ trợ")
                
                service_rate = rates.get(service, {"main": 0, "assistant": 0})
                
                # Main staff payout
                if main_staff and main_staff != "Không cần" and service_rate["main"] > 0:
                    payroll_records.append({
                        "Tháng": f"{month_str}-01",
                        "Nhân sự": main_staff,
                        "Mã hồ sơ": h.get("Mã hồ sơ"),
                        "Công đoạn khoán": f"{service} (Phụ trách chính)",
                        "Số tiền khoán": service_rate["main"],
                        "Phụ cấp": 0,
                        "Thưởng/Phạt": 0,
                        "Tổng nhận": service_rate["main"],
                        "Ngày chốt": datetime.now().strftime("%Y-%m-%d"),
                        "Ghi chú": "Tính lương tự động từ cơ chế khoán"
                    })
                    
                # Assistant staff payout
                if assist_staff and assist_staff != "Không cần" and service_rate["assistant"] > 0:
                    payroll_records.append({
                        "Tháng": f"{month_str}-01",
                        "Nhân sự": assist_staff,
                        "Mã hồ sơ": h.get("Mã hồ sơ"),
                        "Công đoạn khoán": f"{service} (Trợ phụ đo)",
                        "Số tiền khoán": service_rate["assistant"],
                        "Phụ cấp": 0,
                        "Thưởng/Phạt": 0,
                        "Tổng nhận": service_rate["assistant"],
                        "Ngày chốt": datetime.now().strftime("%Y-%m-%d"),
                        "Ghi chú": "Tính lương tự động từ cơ chế khoán phụ"
                    })
                    
    return payroll_records
